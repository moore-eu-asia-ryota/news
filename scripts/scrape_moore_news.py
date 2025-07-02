#!/usr/bin/env python3
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import range_boundaries
from urllib.parse import urljoin
import datetime
import re

BASE_URL     = 'https://www.moore-czech.cz'
LISTING_PATH = '/tiskove-zpravy'
EXCEL_FILE   = 'News.xlsx'
TABLE_NAME   = 'News'
SOURCE_NAME  = 'Moore Czech s.r.o.'

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/115.0.0.0 Safari/537.36'
    )
}

# Czech month abbreviations → month number
MONTH_MAP = {
    'led': 1, 'úno': 2, 'bře': 3, 'dub': 4,
    'kvě': 5, 'čvn': 6, 'čvc': 7, 'srp': 8,
    'zář': 9, 'říj':10, 'lis':11, 'pro':12
}

def parse_czech_date(d: str) -> datetime.date:
    # e.g. "30 čvn 2025"
    day, mon, year = d.split()
    key = mon[:3].lower()
    return datetime.date(int(year), MONTH_MAP[key], int(day))

def main():
    # 1) Load workbook & find table bounds
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Sheet1']
    table = ws.tables[TABLE_NAME]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # 2) Read headers & find URL column
    headers = [ws.cell(min_row, c).value.strip() for c in range(min_col, max_col+1)]
    if 'URL' not in headers:
        raise RuntimeError(f"'URL' column not in headers: {headers}")
    url_idx = headers.index('URL')
    url_col = min_col + url_idx

    # 3) Build set of existing URLs
    existing = {
        ws.cell(r, url_col).value
        for r in range(min_row+1, ws.max_row+1)
        if ws.cell(r, url_col).value
    }

    # 4) GET the listing page
    resp = requests.get(urljoin(BASE_URL, LISTING_PATH), headers=HEADERS)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, 'html.parser')

    # 5) Find the main block (or fallback to body) and all <h5> headings
    main_block = soup.find('div', id='block-system-main') or soup.body
    headings   = main_block.find_all('h5')
    print(f"Found {len(headings)} <h5> headings to scan")

    added = 0
    for h5 in headings:
        a = h5.find('a')
        if not a or not a.get('href', '').startswith('/tiskove-zpravy'):
            continue

        title    = a.get_text(strip=True)
        full_url = urljoin(BASE_URL, a['href'])
        if full_url in existing:
            continue

        # 6) Extract date from the surrounding text
        textblob = h5.parent.get_text(" ", strip=True)
        m = re.search(r'(\d{1,2} \S+ \d{4})', textblob)
        if not m:
            print(f"  ⚠️ couldn’t find date for “{title}”")
            continue
        post_date = parse_czech_date(m.group(1)).strftime('%d.%m.%Y')

        # 7) Scrape the detail page
        r2 = requests.get(full_url, headers=HEADERS)
        r2.raise_for_status()
        s2 = BeautifulSoup(r2.content, 'html.parser')

        body = s2.select_one('.field--name-field-body')
        if body:
            content = body.get_text("\n", strip=True)
        else:
            paras  = s2.select('article p')
            content = "\n\n".join(p.get_text(strip=True) for p in paras)

        # 8) Append the new row
        ws.append([
            title,
            content,
            '',            # Title ENG
            '',            # Summary ENG
            post_date,
            full_url,
            SOURCE_NAME
        ])
        print(f"  ➕ Added: {title}")
        added += 1

    if added:
        wb.save(EXCEL_FILE)
        print(f"Saved workbook with {added} new row(s).")
    else:
        print("No new articles to add.")

if __name__ == '__main__':
    main()
