import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime

# --- CONFIG -------------------------------------------------------
EXCEL_PATH = "PressReleases.xlsx"
SHEET_NAME = "Sheet1"
TABLE_NAME = "PressReleases"     # (Excel table name; we'll adjust its ref)
BASE_LISTING_URL = "https://www.moore-czech.cz/tiskove-zpravy"
MAX_PAGES = 20
# -----------------------------------------------------------------

def strip_tags(text):
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", text)).strip()

def get_seen_urls(ws):
    # Assumes URL column is D (zero-based idx=3) and header in row 1
    seen = set()
    for cell in ws["D"][2:]:   # skip header (row 1) and total row (if any)
        if isinstance(cell.value, str):
            seen.add(cell.value)
    return seen

def fetch_listing():
    new_links = []
    seen = set()
    link_re = re.compile(r'<h5[^>]*>\s*<a\s+href="([^"]+)"', re.IGNORECASE)
    for p in range(1, MAX_PAGES+1):
        resp = requests.get(f"{BASE_LISTING_URL}?page={p}")
        if resp.status_code != 200:
            break
        html = resp.text
        found_new = False
        for m in link_re.finditer(html):
            url = requests.compat.urljoin(BASE_LISTING_URL, m.group(1))
            if url.startswith("https://") and url not in seen:
                seen.add(url)
                new_links.append(url)
                found_new = True
        if not found_new:
            break
    return new_links

def scrape_article(url):
    resp = requests.get(url)
    if resp.status_code != 200:
        return None
    html = resp.text
    # title = first <h1> or fallback <h2>
    m = re.search(r"<h1[^>]*>([\s\S]*?)</h1>", html, re.IGNORECASE) \
        or re.search(r"<h2[^>]*>([\s\S]*?)</h2>", html, re.IGNORECASE)
    title = strip_tags(m.group(1)) if m else None

    # date = first <h4>
    d = re.search(r"<h4[^>]*>([\s\S]*?)</h4>", html, re.IGNORECASE)
    postDate = strip_tags(d.group(1)) if d else None

    # content: after that <h4> up to next share block or next <h4>
    content = ""
    if d:
        idx = html.index(d.group(0)) + len(d.group(0))
        rest = html[idx:]
        share_idx = re.search(r'<ul[^>]+class="share"', rest, re.IGNORECASE)
        chunk = rest[: share_idx.start()] if share_idx else rest.split(r"<h4",1)[0]
        content = strip_tags(chunk)

    if not title:
        return None

    return {
        "Title": title,
        "Content": content,
        "PostDate": postDate,
        "URL": url,
        "Source": "Moore Czech Republic"
    }

def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 1) figure out which URLs we already have
    seen_urls = get_seen_urls(ws)

    # 2) find new press‐release pages
    new_links = [u for u in fetch_listing() if u not in seen_urls]
    if not new_links:
        print("No new releases found.")
        return

    # 3) scrape each one and append a new row
    for link in new_links:
        art = scrape_article(link)
        if not art:
            continue
        # append in the same column order as your table headers
        row = [
            art["Title"],
            art["Content"],
            art["PostDate"],
            art["URL"],
            art["Source"]
        ]
        ws.append(row)

    # 4) (optional) update the Excel table’s ref-range to include new rows
    #    If your table was A1:E10, you can expand it to A1:E{ws.max_row}
    for tbl in ws._tables:
        if tbl.name == TABLE_NAME:
            tbl.ref = f"{tbl.ref.split(':')[0]}:{ 'E' + str(ws.max_row) }"
            break

    # 5) save
    wb.save(EXCEL_PATH)
    print(f"Appended {len(new_links)} new rows.")

if __name__ == "__main__":
    main()
